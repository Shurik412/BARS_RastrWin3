# -*- coding: utf-8 -*-
import csv
import os.path
from os import path, mkdir
import xml.etree.ElementTree as et
from re import findall

import pandas as pd
from openpyxl import (Workbook, load_workbook)
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch, WithEvents


class RastrEvents:
    """
    Метод Onprot - выводит сообщения написанные: rastr.Printp("Сообщение из Printp")\n
    Метод OnLog
    """

    @staticmethod
    def OnLog(code, level, id, name, index, description, formName):
        if code == 2:
            print('[Error]', description)
        elif code == 3:
            print('[Warning]', description)
        elif code == 4:
            print('[Lightbulb]', description)
        elif code == 5:
            print('[Info]', description)
        else:
            print([code, description])

    @staticmethod
    def Onprot(message):
        print(message)


OUTPUT = False
if OUTPUT is True:
    RASTR = Dispatch('Astra.Rastr')
    WithEvents(RASTR, RastrEvents)
else:
    RASTR = Dispatch('Astra.Rastr')

NAME_AREA_DICT = {
    "ti":
        {
            "I1889": "Генерация: ОЭС Центра",
            "I1153": "Потребление: ОЭС Центра",
            "I1494": "Генерация: Владимирское РДУ",
            "I1222": "Потребление: Владимирское РДУ",
            "I1105": "Генерация: Вологодское РДУ",
            "I1107": "Потребление: Вологодское РДУ",
            "I1125": "Генерация: Воронежское РДУ",
            "I1133": "Потребление: Воронежское РДУ",
            "I5576": "Генерация: Костромское РДУ",
            "I5577": "Потребление: Костромское РДУ",
            "I110": "Генерация: ЭС Ивановской обл.",
            "I779": "Потребление: ЭС Ивановской обл.",
            "I280": "Генерация: ЭС Костромской обл.",
            "I3936": "Потребление: ЭС Костромской обл.",
            "I1029": "Генерация: Курское РДУ",
            "I1069": "Потребление: Курское РДУ",
            "I1067": "Генерация: ЭС Белгородской обл.",
            "I1073": "Потребление: ЭС Белгородской обл.",
            "I1079": "Генерация: ЭС Курской обл.",
            "I1080": "Потребление: ЭС Курской обл.",
            "I1086": "Генерация: ЭС Орловской обл.",
            "I1087": "Потребление: ЭС Орловской обл.",
            "I5579": "Генерация: Липецкое РДУ",
            "I5580": "Потребление: Липецкое РДУ",
            "I2603": "Генерация: ЭС Липецкой обл.",
            "I2604": "Потребление: ЭС Липецкой обл.",
            "I126": "Генерация: ЭС Тамбовской обл.",
            "I3229": "Потребление: ЭС Тамбовской обл.",
            "I3641": "Генерация: Московское РДУ",
            "I1299": "Потребление: Московское РДУ",
            "I966": "Генерация: Рязанское РДУ",
            "I1167": "Потребление: Рязанское РДУ",
            "I3580": "Генерация: Смоленское РДУ",
            "I3583": "Потребление: Смоленское РДУ",
            "I112": "Генерация: ЭС Брянской обл.",
            "I757": "Потребление: ЭС Брянской обл.",
            "I3204": "Генерация: ЭС Калужской обл.",
            "I3205": "Потребление: ЭС Калужской обл.",
            "I98": "Генерация: ЭС Смоленской обл.",
            "I380": "Потребление: ЭС Смоленской обл.",
            "I1118": "Генерация: Тверское РДУ",
            "I1150": "Потребление: Тверское РДУ",
            "I1260": "Генерация: Тульское РДУ",
            "I1163": "Потребление: Тульское РДУ",
            "I1052": "Генерация: Ярославское РДУ",
            "I1053": "Потребление: Ярославское РДУ"
        },
    "Excel":
        {
            "Генерация: ОЭС Центра": "Ген_ОЭС_Центра",
            "Потребление: ОЭС Центра": "Пот_ОЭС_Центра",
            "Генерация: Владимирское РДУ": "Ген_Владимир_РДУ",
            "Потребление: Владимирское РДУ": "Пот_Владимир_РДУ",
            "Генерация: Вологодское РДУ": "Ген_Вологод_РДУ",
            "Потребление: Вологодское РДУ": "Пот_Вологод_РДУ",
            "Генерация: Воронежское РДУ": "Ген_Воронеж_РДУ",
            "Потребление: Воронежское РДУ": "Пот_Воронеж_РДУ",
            "Генерация: Костромское РДУ": "Ген_Костром_РДУ",
            "Потребление: Костромское РДУ": "Пот_Костром_РДУ",
            "Генерация: ЭС Ивановской обл.": "Ген_Иванов_обл",
            "Потребление: ЭС Ивановской обл.": "Пот_Иванов_обл",
            "Генерация: ЭС Костромской обл.": "Ген_ЭС_Костр_обл",
            "Потребление: ЭС Костромской обл.": "Пот_ЭС_Костр_обл",
            "Генерация: Курское РДУ": "Ген_Курское_РДУ",
            "Потребление: Курское РДУ": "Пот_Курское_РДУ",
            "Генерация: ЭС Белгородской обл.": "Ген_Белгород_обл",
            "Потребление: ЭС Белгородской обл.": "Пот_Белгород_обл",
            "Генерация: ЭС Курской обл.": "Ген_Курской_обл",
            "Потребление: ЭС Курской обл.": "Пот_Курской_обл",
            "Генерация: ЭС Орловской обл.": "Ген_Орловс_обл",
            "Потребление: ЭС Орловской обл.": "Пот_Орловс_обл",
            "Генерация: Липецкое РДУ": "Ген_Липецкое_РДУ",
            "Потребление: Липецкое РДУ": "Пот_Липецкое_РДУ",
            "Генерация: ЭС Липецкой обл.": "Ген_Липецкой_обл",
            "Потребление: ЭС Липецкой обл.": "Пот_Липецкой_обл",
            "Генерация: ЭС Тамбовской обл.": "Ген_Тамбовск_обл",
            "Потребление: ЭС Тамбовской обл.": "Пот_Тамбовск_обл",
            "Генерация: Московское РДУ": "Ген_Мос_РДУ",
            "Потребление: Московское РДУ": "Пот_Мос_РДУ",
            "Потребление: Рязанское РДУ": "Пот_Рязан_РДУ",
            "Генерация: Смоленское РДУ": "Ген_Смолен_РДУ",
            "Потребление: Смоленское РДУ": "Пот_Смолен_РДУ",
            "Генерация: ЭС Брянской обл.": "Ген_Брянской_обл",
            "Потребление: ЭС Брянской обл.": "Пот_Брянской_обл",
            "Генерация: ЭС Калужской обл.": "Ген_Калужск_обл",
            "Потребление: ЭС Калужской обл.": "Пот_Калужск_обл",
            "Генерация: ЭС Смоленской обл.": "Ген_Смолен_обл",
            "Потребление: ЭС Смоленской обл.": "Пот_Смолен_обл",
            "Генерация: Тверское РДУ": "Ген_Тверское_РДУ",
            "Потребление: Тверское РДУ": "Пот_Тверское_РДУ",
            "Генерация: Тульское РДУ": "Ген_Тульское_РДУ",
            "Потребление: Тульское РДУ": "Пот_Тульское_РДУ",
            "Генерация: Ярославское РДУ": "Ген_Ярослав_РДУ",
            "Потребление: Ярославское РДУ": "Пот_Ярослав_РДУ"
        },
    "Correction":
        {
            "Ген_ОЭС_Центра": "Пот_ОЭС_Центра",
            "Ген_Владимир_РДУ": "Пот_Владимир_РДУ",
            "Ген_Вологод_РДУ": "Пот_Вологод_РДУ",
            "Ген_Воронеж_РДУ": "Пот_Воронеж_РДУ",
            "Ген_Костром_РДУ": "Пот_Костром_РДУ",
            "Ген_Иванов_обл": "Пот_Иванов_обл",
            "Ген_ЭС_Костр_обл": "Пот_ЭС_Костр_обл",
            "Ген_Курское_РДУ": "Пот_Курское_РДУ",
            "Ген_Белгород_обл": "Пот_Белгород_обл",
            "Ген_Курской_обл": "Пот_Курской_обл",
            "Ген_Орловс_обл": "Пот_Орловс_обл",
            "Ген_Липецкое_РДУ": "Пот_Липецкое_РДУ",
            "Ген_Липецкой_обл": "Пот_Липецкой_обл",
            "Ген_Тамбовск_обл": "Пот_Тамбовск_обл",
            "Ген_Мос_РДУ": "Пот_Мос_РДУ",
            "Ген_Смолен_РДУ": "Пот_Смолен_РДУ",
            "Ген_Брянской_обл": "Пот_Брянской_обл",
            "Ген_Калужск_обл": "Пот_Калужск_обл",
            "Ген_Смолен_обл": "Пот_Смолен_обл",
            "Ген_Тверское_РДУ": "Пот_Тверское_РДУ",
            "Ген_Тульское_РДУ": "Пот_Тульское_РДУ",
            "Ген_Ярослав_РДУ": "Пот_Ярослав_РДУ"
        },
    "Correction2":
        {
            "Ген_ОЭС_Центра": "ОЭС_Центра",
            "Ген_Владимир_РДУ": "Владимирск_РДУ",
            "Ген_Вологод_РДУ": "Вологодское_РДУ",
            "Ген_Воронеж_РДУ": "Воронежское_РДУ",
            "Ген_Костром_РДУ": "Костромское_РДУ",
            "Ген_Иванов_обл": "Ивановская_обл",
            "Ген_ЭС_Костр_обл": "ЭС_Костром_обл",
            "Ген_Курское_РДУ": "Курское_РДУ",
            "Ген_Белгород_обл": "Белгородск_обл",
            "Ген_Курской_обл": "Курская_обл",
            "Ген_Орловс_обл": "Орловская_обл",
            "Ген_Липецкое_РДУ": "Липецкое_РДУ",
            "Ген_Липецкой_обл": "Липецкая_обл",
            "Ген_Тамбовск_обл": "Тамбовск_обл",
            "Ген_Мос_РДУ": "Москоское_РДУ",
            "Ген_Смолен_РДУ": "Смоленское_РДУ",
            "Ген_Брянской_обл": "Брянская_обл",
            "Ген_Калужск_обл": "Калужская_обл",
            "Ген_Смолен_обл": "Смоленская_обл",
            "Ген_Тверское_РДУ": "Тверское_РДУ",
            "Ген_Тульское_РДУ": "Тульское_РДУ",
            "Ген_Ярослав_РДУ": "Ярославское_РДУ"

        },
    "Name_area2":
        {
            "ОЭС_Центра": "ОЭС_Центра",
            "Владимирск_РДУ": "Владимирск_РДУ",
            "Вологодское_РДУ": "Вологодское_РДУ",
            "Воронежское_РДУ": "Воронежское_РДУ",
            "Костромское_РДУ": "Костромское_РДУ",
            "Ивановская_обл": "Ивановская_обл",
            "ЭС_Костром_обл": "ЭС_Костром_обл",
            "Курское_РДУ": "Курское_РДУ",
            "Белгородск_обл": "Белгородск_обл",
            "Курская_обл": "Курская_обл",
            "Орловская_обл": "Орловская_обл",
            "Липецкое_РДУ": "Липецкое_РДУ",
            "Липецкая_обл": "Липецкая_обл",
            "Тамбовск_обл": "Тамбовск_обл",
            "Москоское_РДУ": "Москоское_РДУ",
            "Смоленское_РДУ": "Смоленское_РДУ",
            "Брянская_обл": "Брянская_обл",
            "Калужская_обл": "Калужская_обл",
            "Смоленская_обл": "Смоленская_обл",
            "Тверское_РДУ": "Тверское_РДУ",
            "Тульское_РДУ": "Тульское_РДУ",
            "Ярославское_РДУ": "Ярославское_РДУ"

        },
    "npa_name_PDG":
        {
            "10": "МОСКВА",
            "11": "ТУЛА",
            "13": "ИВАНОВО",
            "14": "ВЛАДИМИР",
            "15": "ЯРОСЛАВЛЬ",
            "16": "КОСТРОМА",
            "17": "ТВЕРЬ",
            "18": "СМОЛЕНСК",
            "19": "БРЯНСК",
            "21": "ОРЕЛ",
            "23": "КУРСК",
            "24": "ЛИПЕЦК",
            "26": "ВОРОНЕЖ",
            "27": "БЕЛГОРОД",
            "28": "ТАМБОВ без ПС 220 кВ Давыдовская",
            "29": "ВОЛОГДА",
            "31": "РЯЗАНЬ",
            "32": "КАЛУГА без ПС 220 кВ Метзавод",
            "98": "Потребление ПС 220 кВ Давыдовская",
            "99": "Смежные ЭС"
        },
    "MAG_Terminal_and_BARS":
        {
            "ОЭС_Центра": "-",
            "Владимирск_РДУ": "ВЛАДИМИР",
            "Вологодское_РДУ": "ВОЛОГДА",
            "Воронежское_РДУ": "ВОРОНЕЖ",
            "Костромское_РДУ": "-",
            "Ивановская_обл": "ИВАНОВО",
            "ЭС_Костром_обл": "КОСТРОМА",
            "Курское_РДУ": "-",
            "Белгородск_обл": "БЕЛГОРОД",
            "Курская_обл": "КУРСК",
            "Орловская_обл": "ОРЕЛ",
            "Липецкое_РДУ": "-",
            "Липецкая_обл": "ЛИПЕЦК",
            "Тамбовск_обл": "ТАМБОВбезПС2",
            "Москоское_РДУ": "МОСКВА",
            "Смоленское_РДУ": "-",
            "Смоленская_обл": "СМОЛЕНСК",
            "Брянская_обл": "БРЯНСК",
            "Калужская_обл": "КАЛУГАбезПС2",
            "Тверское_РДУ": "ТВЕРЬ",
            "Тульское_РДУ": "ТУЛА",
            "Ярославское_РДУ": "ЯРОСЛАВЛЬ"
        }
}

mkdir("Files")


def create_dict_out_xml(file_name_xml: str = "file.xml"):
    xroot = et.parse(file_name_xml).getroot()
    df_cols = ["Name", "Time", "Value"]
    data_list = []
    for index, node in enumerate(xroot):
        s_name = node.find("Name").text if node is not None else None
        s_time = node.find("Time").text if node is not None else None
        s_value = node.find("Value").text if node is not None else None
        data_list.append({"Name": s_name, "Time": s_time, "Value": s_value})

    out_df = pd.DataFrame(data_list, columns=df_cols)
    pattern = r'\d{2}.\d{2}.\d{4} \d{2}:\d{2}:\d{2}'

    for key in NAME_AREA_DICT['ti']:
        out_df.Name = out_df.Name.replace(key, NAME_AREA_DICT['ti'][key])

    for value in out_df.Time:
        rename_format_time = findall(pattern, value)
        out_df.Time = out_df.Time.replace(value, rename_format_time[0])

    NAME_FILE_CSV = "Files/csv_area2.csv"
    out_df.to_csv(NAME_FILE_CSV, sep=';', encoding='1251')
    wb = Workbook()
    ws = wb.active
    if os.path.exists(NAME_FILE_CSV):
        print(f'Файл "{NAME_FILE_CSV}"!')
    with open(NAME_FILE_CSV) as file_csv:
        reader = csv.reader(file_csv, delimiter=';')
        for row in reader:
            ws.append(row)
    wb.save("Files/file.xlsx")
    ws = wb.active
    for key in NAME_AREA_DICT['ti']:
        list_name_ = []
        list_time_ = []
        list_value_ = []
        title_ = ''
        for key_ in NAME_AREA_DICT['Excel']:
            if key_ == NAME_AREA_DICT['ti'][key]:
                title_ = NAME_AREA_DICT['Excel'][key_]
        ws_ = wb.create_sheet(title=str(title_))
        for i in range(1, ws.max_row):
            name_excel = ws[f'{get_column_letter(2)}{i}'].value
            if name_excel == NAME_AREA_DICT['ti'][key]:
                _name = ws[f'{get_column_letter(2)}{i}'].value
                _time = ws[f'{get_column_letter(3)}{i}'].value
                _value = ws[f'{get_column_letter(4)}{i}'].value
                list_name_.append(_name)
                list_time_.append(_time)
                list_value_.append(_value)
        for index in range(0, 25):
            ws_[f'{get_column_letter(1)}{index + 2}'] = index
        for index, value in enumerate(list_name_):
            ws_[f'{get_column_letter(2)}{index + 2}'] = value
        for index, value in enumerate(list_time_):
            ws_[f'{get_column_letter(3)}{index + 2}'] = value
        for index, value in enumerate(list_value_):
            ws_[f'{get_column_letter(4)}{index + 2}'] = float(value)
    wb.save("Files/file.xlsx")
    list_name_title_excel = ['Час(точка)', 'Название', 'Время', 'Р, МВт']
    for key in NAME_AREA_DICT['Correction']:
        ws_2 = wb[key]
        ws_1 = wb[NAME_AREA_DICT['Correction'][key]]
        for i, row in enumerate(ws_1.iter_rows()):
            for j, col in enumerate(row):
                ws_2.cell(row=i + 1, column=j + 6).value = col.value
                ws_2.cell(row=1, column=j + 6).value = str(list_name_title_excel[j])
                ws_2.cell(row=1, column=j + 1).value = str(list_name_title_excel[j])
        ws_2.title = NAME_AREA_DICT['Correction2'][key]
    wb.save("Files/file.xlsx")
    for key in NAME_AREA_DICT['Name_area2']:
        _ws = wb[key]
        ch = ScatterChart()
        ch.title = f"Генерация {key}"
        xvalues_ = Reference(_ws,
                             min_col=3,
                             min_row=2,
                             max_row=26)
        values_ = Reference(_ws,
                            min_col=4,
                            min_row=2,
                            max_row=26)
        series_ = Series(values_, xvalues_, title_from_data=False)
        ch.x_axis.title = "Время, час"  # название оси Х
        ch.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch.series.append(series_)
        ch.x_axis.scaling.min = 1
        ch.x_axis.scaling.max = 25
        # ch.y_axis.scaling.min = 0
        # ch.y_axis.scaling.max = 10
        _ws.add_chart(ch, f'{get_column_letter(11)}{str(2)}')

        ch1 = ScatterChart()
        ch1.title = f"Потребление {key}"
        xvalues_ = Reference(_ws,
                             min_col=8,
                             min_row=2,
                             max_row=26)
        values_ = Reference(_ws,
                            min_col=9,
                            min_row=2,
                            max_row=26)
        series_ = Series(values_, xvalues_, title_from_data=False)
        ch1.x_axis.title = "Время, час"  # название оси Х
        ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch1.series.append(series_)
        ch1.x_axis.scaling.min = 1
        ch1.x_axis.scaling.max = 25
        # ch.y_axis.scaling.min = 0
        # ch.y_axis.scaling.max = 10
        _ws.add_chart(ch1, f'{get_column_letter(11)}{str(25)}')
    wb.save("Files/file.xlsx")


def mt_BarsMDP(path_save_excel: str = "Files/file_mpt.xlsx") -> None:
    check_file_mptsmz = os.path.isfile('smzu_mega_XML_UR_MDP.mptsmz')
    check_file_mpt = os.path.isfile('date.mpt')
    if check_file_mpt:
        path_file = 'date.mpt'
    elif check_file_mptsmz:
        path_file = 'smzu_mega_XML_UR_MDP.mptsmz'
    else:
        path_file = None
    SHABLON_NAME = input('SHABLON:'
                         '\n\t- Если "1" то "C:\Program Files\RastrWin3\RastrWin3\SHABLON\мегаточка.mpt";'
                         '\n\t- Если "2" то "D:\BarsMDP\SHABLON\мегаточка_смзу.mpt";'
                         '\n\t- Если "3" то "Без шаблона";'
                         '\n\t- Иначе введите полный путь например '
                         '"C:\Program Files\RastrWin3\RastrWin3\SHABLON\мегаточка.mpt": '
                         '\n\n\t СТРОКА ВВОДА: ')
    if SHABLON_NAME == "1":
        SHABLON = r"C:\Program Files\RastrWin3\RastrWin3\SHABLON\мегаточка.mpt"
        print(f'\nВыбран: {SHABLON}')
    elif SHABLON_NAME == "2":
        SHABLON = r"D:\BarsMDP\SHABLON\мегаточка_смзу.mpt"
        print(f'\nВыбран: {SHABLON}')
    else:
        SHABLON = SHABLON_NAME
        print(f'\nВведен пользователем: {SHABLON}')

    if path_file is not None:
        wb = Workbook()
        RASTR.Load(1, path_file, rf'{SHABLON}')
        for n_point in range(RASTR.GetMinUserPoint(), RASTR.GetMaxUserPoint()):
            wb.create_sheet(title=str(n_point))
            ws = wb[str(n_point)]
            RASTR.ReadPnt(n_point)
            RASTR.rgm("")
            table_area2 = RASTR.Tables("area2")
            list_data_area2 = []
            max_row_table_area2 = table_area2.Count - 1
            for index in range(0, max_row_table_area2):
                name = table_area2.Cols('name').Z(index)
                npa = table_area2.Cols('npa').Z(index)
                pg = table_area2.Cols('pg').Z(index)
                pn = table_area2.Cols('pn').Z(index)
                list_data_area2.append([npa, name, pg, pn])
            for index, name_cell in enumerate(['Номер', 'Название', 'Pген', 'Pнаг']):
                ws[f'{get_column_letter(index + 1)}{1}'] = name_cell
            for index, value in enumerate(list_data_area2):
                ws[f'{get_column_letter(1)}{index + 2}'] = int(value[0])
                ws[f'{get_column_letter(2)}{index + 2}'] = str(value[1])
                ws[f'{get_column_letter(3)}{index + 2}'] = float(value[2])
                ws[f'{get_column_letter(4)}{index + 2}'] = float(value[3])
        wb.save(filename=path_save_excel)


def mt_Excel(path_file: str = "Files/file_mpt.xlsx") -> None:
    wb_new = Workbook()
    wb = load_workbook(filename=path_file)
    ws = wb['25']
    list_ = [x for x in range(25, 48)]
    max_row = ws.max_row
    for row in range(2, max_row):
        for index, sheet_ in enumerate(list_):
            ws_ = wb[str(sheet_)]
            npa_sheet_wb_new = ws_[f'{get_column_letter(1)}{row}'].value
            name_sheet_wb_new = ws_[f'{get_column_letter(2)}{row}'].value
            p_gen_sheet_wb_new = ws_[f'{get_column_letter(3)}{row}'].value
            p_nag_sheet_wb_new = ws_[f'{get_column_letter(4)}{row}'].value

            if index == 0:
                name_sheet_wb_new_ = name_sheet_wb_new[:15].replace(" ", "")
                wb_new.create_sheet(title=name_sheet_wb_new_)
                sheet_wb_new = wb_new[str(name_sheet_wb_new_)]
            else:
                name_sheet_wb_new_ = name_sheet_wb_new[:15].replace(" ", "")
                sheet_wb_new = wb_new[str(name_sheet_wb_new_)]

            sheet_wb_new[f'{get_column_letter(1)}{1}'] = 'Час'
            sheet_wb_new[f'{get_column_letter(2)}{1}'] = 'npa'
            sheet_wb_new[f'{get_column_letter(3)}{1}'] = 'Название'
            sheet_wb_new[f'{get_column_letter(4)}{1}'] = 'Рген'
            sheet_wb_new[f'{get_column_letter(5)}{1}'] = 'Рнаг'

            sheet_wb_new[f'{get_column_letter(1)}{index + 2}'] = int(index + 1)
            sheet_wb_new[f'{get_column_letter(2)}{index + 2}'] = int(npa_sheet_wb_new)
            sheet_wb_new[f'{get_column_letter(3)}{index + 2}'] = str(name_sheet_wb_new)
            sheet_wb_new[f'{get_column_letter(4)}{index + 2}'] = float(p_gen_sheet_wb_new)
            sheet_wb_new[f'{get_column_letter(5)}{index + 2}'] = float(p_nag_sheet_wb_new)

    for sheet_ in wb_new.sheetnames:
        ws = wb_new[str(sheet_)]
        ch = ScatterChart()
        ch.title = f"Генерация"
        xvalues_ = Reference(ws,
                             min_col=1,
                             min_row=2,
                             max_row=24)
        values_ = Reference(ws,
                            min_col=4,
                            min_row=2,
                            max_row=24)
        series_ = Series(values_, xvalues_, title_from_data=False)
        ch.x_axis.title = "Время, час"  # название оси Х
        ch.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch.series.append(series_)
        ch.x_axis.scaling.min = 1
        ch.x_axis.scaling.max = 23
        # ch.y_axis.scaling.min = 0
        # ch.y_axis.scaling.max = 10
        ws.add_chart(ch, f'{get_column_letter(7)}{str(2)}')

        ch1 = ScatterChart()
        ch1.title = f"Потребление"
        xvalues_ = Reference(ws,
                             min_col=1,
                             min_row=2,
                             max_row=24)
        values_ = Reference(ws,
                            min_col=5,
                            min_row=2,
                            max_row=24)
        series_ = Series(values_, xvalues_, title_from_data=False)
        ch1.x_axis.title = "Время, час"  # название оси Х
        ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch1.series.append(series_)
        ch1.x_axis.scaling.min = 1
        ch1.x_axis.scaling.max = 23
        # ch.y_axis.scaling.min = 0
        # ch.y_axis.scaling.max = 10
        ws.add_chart(ch1, f'{get_column_letter(7)}{str(17)}')
    wb_new.save('Files/file_mpt_excel.xlsx')


def compare_excel_and_mpt() -> None:
    wb1 = load_workbook('Files/file.xlsx')
    wb2 = load_workbook('Files/file_mpt_excel.xlsx')
    for index, name_sheet in enumerate(NAME_AREA_DICT['MAG_Terminal_and_BARS']):
        ws1 = wb1[str(name_sheet)]
        basic_list = []
        try:
            ws2 = wb2[str(NAME_AREA_DICT['MAG_Terminal_and_BARS'][str(name_sheet)])]
            for i_ in range(1, 26):
                list_ = []
                for j_ in range(1, 10):
                    value = ws1[f'{get_column_letter(j_)}{i_}'].value
                    list_.append(value)
                basic_list.append(list_)
            for index_i, i_ in enumerate(basic_list):
                for index_j, j_ in enumerate(basic_list[index_i]):
                    ws2[f'{get_column_letter(index_j + 1)}{index_i + 35}'].value = j_
        except KeyError:
            print(f'\tKeyError: "{name_sheet}" - "{str(NAME_AREA_DICT["MAG_Terminal_and_BARS"][str(name_sheet)])}"')

    for index, name_sheet in enumerate(wb2.sheetnames):
        ws = wb2[str(name_sheet)]
        if name_sheet != 'Sheet':
            ch1 = ScatterChart()
            ch1.title = f"Генерация"
            xvalues_ = Reference(ws,
                                 min_col=1,
                                 min_row=2,
                                 max_row=24)
            values_ = Reference(ws,
                                min_col=4,
                                min_row=2,
                                max_row=24)
            series_ = Series(values_, xvalues_, title_from_data=False)
            ch1.x_axis.title = "Время, час"  # название оси Х
            ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
            ch1.series.append(series_)

            xvalues_1 = Reference(ws,
                                  min_col=1,
                                  min_row=37,
                                  max_row=59)

            values_1 = Reference(ws,
                                 min_col=4,
                                 min_row=37,
                                 max_row=59)
            series_1 = Series(values_1, xvalues_1, title_from_data=False)
            ch1.series.append(series_1)
            ch1.x_axis.scaling.min = 1
            ch1.x_axis.scaling.max = 23
            # ch.y_axis.scaling.min = 0
            # ch.y_axis.scaling.max = 10
            ws.add_chart(ch1, f'{get_column_letter(12)}{str(35)}')

            ch2 = ScatterChart()
            ch2.title = f"Потребление"
            xvalues_ = Reference(ws,
                                 min_col=1,
                                 min_row=2,
                                 max_row=24)
            values_ = Reference(ws,
                                min_col=5,
                                min_row=2,
                                max_row=24)
            series_ = Series(values_, xvalues_, title_from_data=False)
            ch2.x_axis.title = "Время, час"  # название оси Х
            ch2.y_axis.title = "Активная мощность, МВт"  # название оси У
            ch2.series.append(series_)

            xvalues_1 = Reference(ws,
                                  min_col=1,
                                  min_row=37,
                                  max_row=59)

            values_1 = Reference(ws,
                                 min_col=9,
                                 min_row=37,
                                 max_row=59)
            series_1 = Series(values_1, xvalues_1, title_from_data=False)
            ch2.series.append(series_1)
            ch2.x_axis.scaling.min = 1
            ch2.x_axis.scaling.max = 23
            # ch.y_axis.scaling.min = 0
            # ch.y_axis.scaling.max = 10
            ws.add_chart(ch2, f'{get_column_letter(12)}{str(50)}')

    wb2.save('Files/MAG_MPT_Excel.xlsx')
    print("\nСохранен файл: Files/MAG_MPT_Excel.xlsx")


def file_output() -> None:
    wb_new = Workbook()
    wb_old = load_workbook(filename='Files/MAG_MPT_Excel.xlsx')
    list_name_sheet_old = []
    for index, name_sheet in enumerate(wb_old.sheetnames):
        list_name_sheet_old.append(name_sheet)
    list_name_sheet_old.remove('Sheet')
    ws_new = wb_new.active
    ws_new.title = 'Графики'
    for name_sheet in list_name_sheet_old:
        wb_new.create_sheet(title=name_sheet)
    for name_sheet in wb_new.sheetnames:
        if name_sheet != 'Графики':
            ws_old = wb_old[str(name_sheet)]
            ws_new = wb_new[str(name_sheet)]
            for index_row, row in enumerate(ws_old.iter_rows(min_row=1, max_col=5, max_row=24, values_only=True)):
                for index_column in range(0, 5):
                    ws_new[f'{get_column_letter(index_column + 1)}{index_row + 2}'].value = row[index_column]
                ws_new[f'{get_column_letter(1)}{1}'].value = 'Мегаточка'

            for index_row, row in enumerate(ws_old.iter_rows(min_row=35, max_col=9, max_row=60, values_only=True)):
                for index_column in range(0, 9):
                    ws_new[f'{get_column_letter(index_column + 8)}{index_row + 2}'].value = row[index_column]
                ws_new[f'{get_column_letter(8)}{1}'].value = 'ОИК'

            ch1 = ScatterChart()
            ch1.title = f"Генерация"
            ch1.x_axis.title = "Время, час"  # название оси Х
            ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
            ch1.x_axis.scaling.min = 1
            ch1.x_axis.scaling.max = 23
            # ch.y_axis.scaling.min = 0
            # ch.y_axis.scaling.max = 10

            xvalues_ = Reference(ws_new,
                                 min_col=1,
                                 min_row=3,
                                 max_row=25)
            values_ = Reference(ws_new,
                                min_col=4,
                                min_row=3,
                                max_row=25)
            series_ = Series(values_, xvalues_, title_from_data=False, title="Мегаточка")
            ch1.series.append(series_)

            xvalues_1 = Reference(ws_new,
                                  min_col=8,
                                  min_row=3,
                                  max_row=26)

            values_1 = Reference(ws_new,
                                 min_col=11,
                                 min_row=3,
                                 max_row=26)
            series_1 = Series(values_1, xvalues_1, title_from_data=False, title='ОИК')
            ch1.series.append(series_1)
            ws_new.add_chart(ch1, f'{get_column_letter(1)}{str(28)}')

            ch2 = ScatterChart()
            ch2.title = f"Потребление"
            ch2.x_axis.title = "Время, час"  # название оси Х
            ch2.y_axis.title = "Активная мощность, МВт"  # название оси У
            ch2.x_axis.scaling.min = 1
            ch2.x_axis.scaling.max = 23
            # ch.y_axis.scaling.min = 0
            # ch.y_axis.scaling.max = 10

            xvalues_ = Reference(ws_new,
                                 min_col=1,
                                 min_row=3,
                                 max_row=25)
            values_ = Reference(ws_new,
                                min_col=5,
                                min_row=3,
                                max_row=25)
            series_ = Series(values_, xvalues_, title_from_data=False, title="Мегаточка")
            ch2.series.append(series_)

            xvalues_1 = Reference(ws_new,
                                  min_col=13,
                                  min_row=3,
                                  max_row=26)

            values_1 = Reference(ws_new,
                                 min_col=16,
                                 min_row=3,
                                 max_row=26)
            series_1 = Series(values_1, xvalues_1, title_from_data=False, title='ОИК')
            ch2.series.append(series_1)
            ws_new.add_chart(ch2, f'{get_column_letter(11)}{str(28)}')
    wb_new.save(filename='Output_Excel.xlsx')


def main():
    create_dict_out_xml()
    mt_BarsMDP()
    mt_Excel()
    compare_excel_and_mpt()
    file_output()


main()
# input('\tНажмите Enter!')
