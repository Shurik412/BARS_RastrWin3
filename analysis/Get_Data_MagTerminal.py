# -*- coding: utf-8 -*-
import base64
import datetime
from http.client import HTTPSConnection
import json

from openpyxl import Workbook

## Аутентификация
login = b"Ohrimenko_AG"  ## Логин
password = b"Shurik4122"  ## Пароль
headers = {"Authorization": b"Basic " + base64.b64encode(login + b":" + password)}
connection = HTTPSConnection("cn-ck11-web-ep.oducn.so:9443")
connection.request("POST", "/auth/app/token", "", headers)
response = connection.getresponse()
# print("Status: {} and reason: {}".format(response.status, response.reason))

responceBody = response.read()
responseData = json.loads(responceBody)
# print(responseData)
headers = {
    "Authorization": responseData["token_type"] + " " + responseData["access_token"],
    "Content-Type": "application/json"
}

# Подключение и получение данных
connection = HTTPSConnection("cn-ck11-web-ep.oducn.so")
# POST запрос
payload = json.dumps(
    {
        "uids": ["a4343a74-d12e-4f83-93dc-8e0ca652962e"],
        "fromTimeStamp": "2022-01-25T00:00:00.000Z",
        "toTimeStamp": "2022-01-25T23:59:00.000Z",
        "stepUnits": "seconds",
        "stepValue": 60
    }
)
connection.request("POST", "/api/public/measurement-values/v2.0/numeric/data/get-table", payload, headers)
response = connection.getresponse()
print("Status: {} and reason: {}".format(response.status, response.reason))
# print(response.read())

data_json = json.loads(response.read())


def date_remove(s):
    s = s.replace("Z", "")
    s = s.replace("T", " ")
    date_time_str = s
    try:
        date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
    # print('Date:', date_time_obj.date())
    # print('Time:', date_time_obj.time())
    # print('Date-time:', date_time_obj)
    date = date_time_obj.date()
    time = date_time_obj.time()
    date_time = date_time_obj
    return date, time, date_time


wb = Workbook()
ws = wb.active
ws[f'A{1}'].value = 'Время'
ws[f'B{1}'].value = 'Значения'

for key, i in enumerate(data_json["value"][0]['value']):
    print(f'{key}. {date_remove(i["timeStamp"])[2]} -> {i["value"]}')
    ws[f'A{key + 2}'].value = date_remove(i["timeStamp"])[2]
    ws[f'B{key + 2}'].value = i["value"]

connection.close()

wb.save(r'L:\SER\Okhrimenko\19.СК-11(АИП)\Excel2201.xlsx')


