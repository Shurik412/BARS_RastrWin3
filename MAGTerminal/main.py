# -*- coding: utf-8 -*-
import base64
import datetime
import json
import time
from http.client import HTTPSConnection
from json import dumps

from openpyxl import Workbook
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.utils import get_column_letter


def post_json(uid: str, date_1: str, date_2: str, step_seconds: int = 1) -> str:
    """

    :param uid: UID - индентификационный номер
    :param date_1: 2022-01-25T10:05:00.000Z
    :param date_2: 2022-01-25T20:05:00.000Z
    :param step_seconds: шаг в секундах
    :return:
    """
    payload = dumps(
        {
            "uids": [uid],
            "fromTimeStamp": date_1,
            "toTimeStamp": date_2,
            "stepUnits": "seconds",
            "stepValue": step_seconds
        }
    )
    return payload


def changing_number_of_semicolons(number, digits=0):
    """ """
    return f"{number:.{digits}f}"


def date_remove(date_str: str) -> tuple:
    """

    :param date_str:
    :return:
    """
    date_str = date_str.replace("Z", "")
    date_str = date_str.replace("T", " ")
    try:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    # print('Date:', date_time_obj.date())
    # print('Time:', date_time_obj.time())
    # print('Date-time:', date_time_obj)
    date_ = date_time_obj.date()
    time_ = date_time_obj.time()
    return date_, time_, date_time_obj


def date_remove_csv(date_str: str) -> tuple:
    """

    :param date_str:
    :return:
    """
    try:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    # print('Date:', date_time_obj.date())
    # print('Time:', date_time_obj.time())
    # print('Date-time:', date_time_obj)
    date_ = date_time_obj.date()
    time_ = date_time_obj.time()
    return date_, time_, date_time_obj


def get_data(uid: str, date_1: str, date_2: str, step_seconds: int):
    """

    :param uid:
    :param date_1: "2022-01-25T00:10:00.000Z"
    :param date_2: "2022-01-25T23:10:00.000Z"
    :param step_seconds: 30 секунд
    :return: dict
    """
    login_ = b"Ohrimenko_AG"
    password_ = b"Shurik4122"

    headers = {"Authorization": b"Basic " + base64.b64encode(login_ + b":" + password_)}
    connection = HTTPSConnection("cn-ck11-web-ep.oducn.so:9443")
    connection.request(
        "POST",
        "/auth/app/token",
        "",
        headers
    )
    response = connection.getresponse()
    responceBody = response.read()
    responseData = json.loads(responceBody)

    headers = {
        "Authorization": responseData["token_type"] + " " + responseData["access_token"],
        "Content-Type": "application/json"
    }

    # Подключение и получение данных
    connection = HTTPSConnection("cn-ck11-web-ep.oducn.so")
    payload = post_json(uid=uid,
                        date_1=date_1,
                        date_2=date_2,
                        step_seconds=step_seconds)

    connection.request(
        "POST",
        "/api/public/measurement-values/v2.0/numeric/data/get-table",
        payload,
        headers
    )
    response = connection.getresponse()
    print(f"Status: {response.status} and reason: {response.reason}")
    data_json = json.loads(response.read())
    connection.close()

    return data_json


def remove_csv(date: str):
    """

    :param date: "2022-01-24"
    :return:
    """
    # date = "2022-01-24"

    name_sech_dict = {
        "10": "КС Донское (ток)",
        "19": "КС Донское (статика)",
        "2": "КС Воронежское-2 на Север (ток)",
        "18": "КС Воронежское-2 на Север (статика)",
    }

    date_dict = {
        "25": f'{date} 00:00:00',
        "26": f'{date} 01:00:00',
        "27": f'{date} 02:00:00',
        "28": f'{date} 03:00:00',
        "29": f'{date} 04:00:00',
        "30": f'{date} 05:00:00',
        "31": f'{date} 06:00:00',
        "32": f'{date} 07:00:00',
        "33": f'{date} 08:00:00',
        "34": f'{date} 09:00:00',
        "35": f'{date} 10:00:00',
        "36": f'{date} 11:00:00',
        "37": f'{date} 12:00:00',
        "38": f'{date} 13:00:00',
        "39": f'{date} 14:00:00',
        "40": f'{date} 15:00:00',
        "41": f'{date} 16:00:00',
        "42": f'{date} 17:00:00',
        "43": f'{date} 18:00:00',
        "44": f'{date} 19:00:00',
        "45": f'{date} 20:00:00',
        "46": f'{date} 21:00:00',
        "47": f'{date} 22:00:00',
        "48": f'{date} 23:00:00',
    }

    list_ = []
    file_BARS_MDP = 'ogrsech.csv'
    with open(file_BARS_MDP) as file:
        file_line = file.readlines()
        for i, j in enumerate(file_line):
            name_sech = j.split(";")[0].replace(j.split(";")[0], name_sech_dict[j.split(";")[0]])
            t = (j.split(";")[1].replace(j.split(";")[1], date_dict[j.split(";")[1]]))
            list_.append([name_sech, t, j.split(";")[2]])
            # print(j.split(";")[0].replace(j.split(";")[0], name_sech_dict[j.split(";")[0]]))
    return list_


def delta_datetime_hours(date, delta):
    date += datetime.timedelta(hours=delta)
    return date


def max_row_bars_data(ws, name_cols: str = "L") -> int:
    k = 0
    for i in range(2, ws.max_row):
        if ws[f'{name_cols}{i}'].value is None:
            k = i - 1
            break
    return k


def get_result(name_sech: str, uid_: list, date_: str):
    """

    :param date_:
    :param name_sech:
    :param uid_:
    :return:
    """
    wb = Workbook()
    ws = wb.active

    print(f"\n{name_sech}")

    ws['A1'].value = name_sech

    # date_ = "2022-01-18"

    for key_, uid in enumerate(uid_):
        list1 = get_data(
            uid=uid[0],
            date_1=f'{date_}T00:00:00.000Z',
            date_2=f'{date_}T23:59:59.000Z',
            step_seconds=10
        )
        print(f"{key_}. {uid[0]} - {uid[1]}")
        if key_ == 0:
            ws[f'{get_column_letter(key_ + 1)}2'].value = 'Время'
            ws[f'{get_column_letter(key_ + 2)}2'].value = uid[1]
        else:
            ws[f'{get_column_letter(key_ * 2 + 1)}2'].value = 'Время'
            ws[f'{get_column_letter(key_ * 2 + 2)}2'].value = uid[1]

        for key, value in enumerate(list1['value'][0]['value']):
            ws[f"{get_column_letter(key_ * 2 + 1)}{key + 3}"].value = date_remove(value['timeStamp'])[2]
            ws[f"{get_column_letter(key_ * 2 + 2)}{key + 3}"].value = value['value']

    list_csv = remove_csv(date=date_)
    f = 0
    i = 0
    for key_bars, value_bars in enumerate(list_csv):
        if name_sech == 'КС Донское':
            name_sech_csv = value_bars[0]
            if name_sech_csv == 'КС Донское (ток)':
                f = f + 2
                if f == 2:
                    ws['L2'].value = "Время"
                    ws['M2'].value = "КС Донское (ток)"

                ws[f'L{f + 2}'].value = date_remove(date_str=value_bars[1])[2]
                ws[f'M{f + 2}'].value = float(value_bars[2])

            if name_sech_csv == 'КС Донское (статика)':
                i = i + 2
                if i == 2:
                    ws['N2'].value = "Время"
                    ws['O2'].value = "КС Донское (статика)"

                ws[f'N{i + 2}'].value = date_remove(date_str=value_bars[1])[2]
                ws[f'O{i + 2}'].value = float(value_bars[2])

        elif name_sech == 'КС Воронежское-2 на Север':
            name_sech_csv = value_bars[0]
            if name_sech_csv == 'КС Воронежское-2 на Север (ток)':
                f = f + 2
                if f == 2:
                    ws['L2'].value = "Время"
                    ws['M2'].value = "КС Воронежское-2 на Север (ток)"

                ws[f'L{f + 2}'].value = date_remove(date_str=value_bars[1])[2]
                ws[f'M{f + 2}'].value = float(value_bars[2])

            if name_sech_csv == 'КС Воронежское-2 на Север (статика)':
                i = i + 2
                if i == 2:
                    ws['N2'].value = "Время"
                    ws['O2'].value = "КС Воронежское-2 на Север (статика)"

                ws[f'N{i + 2}'].value = date_remove(date_str=value_bars[1])[2]
                ws[f'O{i + 2}'].value = float(value_bars[2])
        else:
            print("name_sech - не найдено!")

    for i in range(0, 24):
        time_L_copy = ws[f'L{i * 2 + 4}'].value
        value_M_copy = ws[f'M{i * 2 + 4}'].value
        time_N_copy = ws[f'N{i * 2 + 4}'].value
        value_O_copy = ws[f'O{i * 2 + 4}'].value

        ws[f'L{i * 2 + 3}'].value = time_L_copy
        ws[f'M{i * 2 + 3}'].value = value_M_copy
        ws[f'N{i * 2 + 3}'].value = time_N_copy
        ws[f'O{i * 2 + 3}'].value = value_O_copy
        if i == 23:
            time_L_before = delta_datetime_hours(
                date=time_L_copy,
                delta=0.99
            )
            ws[f'L{i * 2 + 4}'].value = time_L_before

            time_N_before = delta_datetime_hours(
                date=time_N_copy,
                delta=0.99
            )
            ws[f'N{i * 2 + 4}'].value = time_N_before
        else:
            time_L_before = delta_datetime_hours(
                date=time_L_copy,
                delta=1
            )
            ws[f'L{i * 2 + 4}'].value = time_L_before

            time_N_before = delta_datetime_hours(
                date=time_N_copy,
                delta=1
            )
            ws[f'N{i * 2 + 4}'].value = time_N_before

    ws[f'Q2'].value = f'Время'
    ws[f'R2'].value = f'{name_sech} БАРС-МДП'

    for i in range(3, 51):
        time_Q_copy = ws[f'L{i}'].value
        ws[f'Q{i}'].value = time_Q_copy

        value_M_current = ws[f'M{i}'].value
        value_O_statics = ws[f'O{i}'].value
        # print(type(min(value_M_current, value_O_statics)), min(value_M_current, value_O_statics))
        ws[f'R{i}'].value = min(value_M_current, value_O_statics)

    ch1 = ScatterChart()
    xvalues_ = Reference(ws,
                         min_col=1,
                         min_row=3,
                         max_row=ws.max_row)

    values_ = Reference(ws,
                        min_col=2,
                        min_row=2,
                        max_row=ws.max_row)

    xvalues_2 = Reference(ws,
                          min_col=17,
                          min_row=3,
                          max_row=50)

    values_2 = Reference(ws,
                         min_col=18,
                         min_row=2,
                         max_row=50)

    # xvalues_3 = Reference(ws,
    #                       min_col=14,
    #                       min_row=3,
    #                       max_row=50)
    #
    # values_3 = Reference(ws,
    #                      min_col=15,
    #                      min_row=2,
    #                      max_row=50)

    series_1 = Series(values_, xvalues_, title_from_data=True)
    series_2 = Series(values_2, xvalues_2, title_from_data=True)
    # series_3 = Series(values_3, xvalues_3, title_from_data=True)

    ch1.title = f"{name_sech}"
    ch1.x_axis.title = "Время, с"  # название оси Х
    ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
    ch1.legend.position = 'b'

    ch1.series.append(series_1)
    ch1.series.append(series_2)
    # ch1.series.append(series_3)

    # ch1.x_axis.scaling.min = 49
    # ch1.y_axis.scaling.min = 0
    # ch1.y_axis.scaling.max = max_Chart + 10
    # ch1.x_axis.scaling.max = 65

    ws.add_chart(ch1, f'{get_column_letter(20)}{str(3)}')
    dir_file = rf'{name_sech}.xlsx'
    try:
        wb.save(dir_file)
    except PermissionError:
        print(f'Не сохранен файл: {dir_file}')


name_sech_don = 'КС Донское'
uid_don = [
    ['a4343a74-d12e-4f83-93dc-8e0ca652962e', 'МДП с ПА'],
    ['fbb8dce0-9a4c-48da-b16a-374a28933570', 'ток'],
    ['93e9420f-d2bb-4c99-86c2-64e2b68eab7e', 'статика'],
    ['d8ad2327-0645-44e6-b6b7-38f4fd505314', 'ПУР'],
    ['18dc8b42-c075-44e9-bb02-a251f28df8c0', 'Фактическое значение'],
]

name_sech_vor = 'КС Воронежское-2 на Север'
uid_vor = [
    ['5332078b-9940-425c-8072-541181c8c005', 'МДП с ПА'],
    ['468b9602-6701-4bdf-ab7c-9b5b5796d446', 'ток'],
    ['413b29fc-c9a7-4255-82f9-bd1e20b5fded', 'статика'],
    ['2a662415-4e1c-46db-b78d-79bb41134ba9', 'ПУР'],
    ['138d81e2-eaa4-4647-ac23-5907553cafed', 'Фактическое значение'],
]

start_time = time.time()
date_input = input("Введите дату (пример: 2022-01-18): ")
if date_input != "":
    get_result(name_sech=name_sech_don, uid_=uid_don, date_=date_input)
    get_result(name_sech=name_sech_vor, uid_=uid_vor, date_=date_input)
else:
    print('Выход. Вы не указали дату!')
print(f'Время работы макроса:')
print(f"--- {changing_number_of_semicolons((time.time() - start_time), digits=0)} сек. ---")
print("The End.")
print("---------------------")
input('Нажмите клавишу Enter')
