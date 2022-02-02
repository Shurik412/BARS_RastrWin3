# -*- coding: utf-8 -*-
import datetime


def date_remove(date_str: str) -> tuple:
    """

    :param date_str:
    :return:
    """
    date_str = date_str.replace("Z", "")
    date_str = date_str.replace("T", " ")
    try:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S.%f')
    except ValueError:
        date_time_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
    # print('Date:', date_time_obj.date())
    # print('Time:', date_time_obj.time())
    # print('Date-time:', date_time_obj)
    date = date_time_obj.date()
    time = date_time_obj.time()
    return date, time, date_time_obj


if __name__ == '__main__':
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(filename=r'L:\SER\Okhrimenko\Project_Py3\BARS_RastrWin3\MAGTerminal\КС Воронежское-2 на Север.xlsx')
    ws = wb.active

    for i in range(2, ws.max_row):
        print(ws[f'L{i}'].value)
        if ws[f'L{i}'].value is None:
            k = i - 1
            break
        else:
            k = 0
    print(k)

