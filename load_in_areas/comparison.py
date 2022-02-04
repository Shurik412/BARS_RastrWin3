# -*- coding: utf-8 -*-
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch

from .moduls.AstraRastr import RASTR
from .moduls.Load import LoadFile
from .moduls.Templates import changing_number_of_semicolons
from .moduls.get import GettingParameter


def main(name_file_mpt_bars: str) -> None:
    """
    Нужный файлы:
     1. smzu_mega_XML_UR_MDP.mptsmz
     2. pload.csv
     3. pgen.csv
     4. name_file_mpt_bars -> 180222-17.mpt
    :return: Nothing
    """

    nameFileMPT = f'{name_file_mpt_bars}-17.mpt'

    area2 = {
        "742": "Потребление ПС 220 кВ Метзавод",
        "980": "Потребление ПС 220 кВ Давыдовская",
        "982": "БЕЛГОРОД",
        "983": "БРЯНСК",
        "984": "ВЛАДИМИР",
        "985": "ВОЛОГДА",
        "986": "ВОРОНЕЖ",
        "987": "ИВАНОВО",
        "988": "КАЛУГА (без учета 5742 ПС 220 кВ Метзавод)",
        "989": "КОСТРОМА",
        "990": "КУРСК",
        "991": "ЛИПЕЦК",
        "992": "МОСКВА",
        "993": "ОРЕЛ",
        "994": "РЯЗАНЬ",
        "995": "СМОЛЕНСК",
        "996": "ТАМБОВ (без учета 5980 ПС 220 кВ Давыдовская)",
        "997": "ТВЕРЬ",
        "998": "ТУЛА",
        "999": "ЯРОСЛАВЛЬ"
    }
    list_area2 = [
        "742", "980", "982", "983", "984",
        "985", "986", "987", "988", "989",
        "990", "991", "992", "993", "994",
        "995", "996", "997", "998", "999",
    ]

    load_obj = LoadFile(rastr_win=RASTR)
    load_obj.load(path_file='smzu_mega_XML_UR_MDP.mptsmz', name_shabl_russian='мегаточка')

    print(f"Загружен файл: smzu_mega_XML_UR_MDP.mptsmz")

    get = GettingParameter(rastr_win=RASTR)

    wb = Workbook()
    ws = wb.active

    print(f"Загружен файл: pload.csv")
    csv_ = pd.read_csv('pload.csv', delimiter=';', header=None)
    headerCSV = ['Num', 'Point', 'P']
    csv_.columns = headerCSV
    ws['A1'].value = 'БАРС - ПДГ'
    step = 3
    step2 = 2
    col = 0
    for area in area2:
        dd = csv_[csv_.Num == int(area)].to_dict()
        ws[f'{get_column_letter(col + 1)}{2}'].value = area2[area]
        for j in headerCSV:
            row = 0
            row2 = 0
            col = col + 1
            ws[f'{get_column_letter(col)}{step}'].value = f'{j} - {area2[area]}'
            for i in dd[j]:
                row += 2
                row2 += 2
                ws[f'{get_column_letter(col)}{row2 + step2}'].value = dd[j][i]
                if j == "Point":
                    ws[f'{get_column_letter(col)}{row + step}'].value = dd[j][i] + 1
                else:
                    ws[f'{get_column_letter(col)}{row + step}'].value = dd[j][i]
    print(f"Выгружены данные из файла: pload.csv")
    area2 = RASTR.Tables("area2")
    POINT_START = 25
    POINT_END = 49
    row_start = 55
    row = 0
    h = 0
    for point in range(POINT_START, POINT_END):
        RASTR.ReadPnt(point)
        RASTR.rgm("")
        row = row + 1
        col = 1
        kk = 5
        h += 1
        for key, area in enumerate(list_area2):
            area2.SetSel(f"npa={area}")
            row_id = area2.FindNextSel(-1)
            if row_id != (-1):
                npa = get.get_cell_row(table="area2", column="npa", row=row_id)
                name = get.get_cell_row(table="area2", column="name", row=row_id)
                pn = get.get_cell_row(table="area2", column="pn", row=row_id)
                pg = get.get_cell_row(table="area2", column="pg", row=row_id)
                pop = get.get_cell_row(table="area2", column="pop", row=row_id)
                if row == 1:
                    if col == 1:
                        ws[f'{get_column_letter(col)}{row_start}'].value = 'Nр-н'
                        ws[f'{get_column_letter(col)}{row_start + row + h}'].value = npa

                        ws[f'{get_column_letter(col)}{row_start - 1 + row + h}'].value = npa

                        ws[f'{get_column_letter(col + 1)}{row_start}'].value = 'Point'
                        ws[f'{get_column_letter(col + 1)}{row_start + row + h}'].value = point + 1

                        ws[f'{get_column_letter(col + 1)}{row_start - 1 + row + h}'].value = point

                        ws[f'{get_column_letter(col + 2)}{row_start}'].value = 'Район'
                        ws[f'{get_column_letter(col + 2)}{row_start + row + h}'].value = name

                        ws[f'{get_column_letter(col + 2)}{row_start - 1 + row + h}'].value = name

                        ws[f'{get_column_letter(col + 3)}{row_start}'].value = f'Pнаг {name}'
                        ws[f'{get_column_letter(col + 3)}{row_start + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 3)}{row_start - 1 + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 4)}{row_start}'].value = f'Pпотр {name}'
                        ws[f'{get_column_letter(col + 4)}{row_start + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 4)}{row_start - 1 + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 5)}{row_start}'].value = f'Pген {name}'
                        ws[f'{get_column_letter(col + 5)}{row_start + row + h}'].value = pg

                        ws[f'{get_column_letter(col + 5)}{row_start - 1 + row + h}'].value = pg
                    else:
                        ws[f'{get_column_letter(col + kk)}{row_start}'].value = 'Nр-н'
                        ws[f'{get_column_letter(col + kk)}{row_start + row + h}'].value = npa

                        ws[f'{get_column_letter(col + kk)}{row_start - 1 + row + h}'].value = npa

                        ws[f'{get_column_letter(col + 1 + kk)}{row_start}'].value = 'Point'
                        ws[f'{get_column_letter(col + 1 + kk)}{row_start + row + h}'].value = point + 1

                        ws[f'{get_column_letter(col + 1 + kk)}{row_start - 1 + row + h}'].value = point

                        ws[f'{get_column_letter(col + 2 + kk)}{row_start}'].value = 'Район'
                        ws[f'{get_column_letter(col + 2 + kk)}{row_start + row + h}'].value = name

                        ws[f'{get_column_letter(col + 2 + kk)}{row_start - 1 + row + h}'].value = name

                        ws[f'{get_column_letter(col + 3 + kk)}{row_start}'].value = f'Pнаг {name}'
                        ws[f'{get_column_letter(col + 3 + kk)}{row_start + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 3 + kk)}{row_start - 1 + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 4 + kk)}{row_start}'].value = f'Pпотр {name}'
                        ws[f'{get_column_letter(col + 4 + kk)}{row_start + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 4 + kk)}{row_start - 1 + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 5 + kk)}{row_start}'].value = f'Pген {name}'
                        ws[f'{get_column_letter(col + 5 + kk)}{row_start + row + h}'].value = pg

                        ws[f'{get_column_letter(col + 5 + kk)}{row_start - 1 + row + h}'].value = pg

                        kk += 5
                else:
                    if col == 1:
                        ws[f'{get_column_letter(col)}{row_start}'].value = 'Nр-н'
                        ws[f'{get_column_letter(col)}{row_start + row + h}'].value = npa

                        ws[f'{get_column_letter(col)}{row_start - 1 + row + h}'].value = npa

                        ws[f'{get_column_letter(col + 1)}{row_start}'].value = 'Point'
                        ws[f'{get_column_letter(col + 1)}{row_start + row + h}'].value = point + 1

                        ws[f'{get_column_letter(col + 1)}{row_start - 1 + row + h}'].value = point

                        ws[f'{get_column_letter(col + 2)}{row_start}'].value = 'Район'
                        ws[f'{get_column_letter(col + 2)}{row_start + row + h}'].value = name

                        ws[f'{get_column_letter(col + 2)}{row_start - 1 + row + h}'].value = name

                        ws[f'{get_column_letter(col + 3)}{row_start}'].value = f'Pнаг {name}'
                        ws[f'{get_column_letter(col + 3)}{row_start + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 3)}{row_start - 1 + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 4)}{row_start}'].value = f'Pпотр {name}'
                        ws[f'{get_column_letter(col + 4)}{row_start + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 4)}{row_start - 1 + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 5)}{row_start}'].value = f'Pген {name}'
                        ws[f'{get_column_letter(col + 5)}{row_start + row + h}'].value = pg

                        ws[f'{get_column_letter(col + 5)}{row_start - 1 + row + h}'].value = pg
                    else:
                        ws[f'{get_column_letter(col + kk)}{row_start}'].value = 'Nр-н'
                        ws[f'{get_column_letter(col + kk)}{row_start + row + h}'].value = npa

                        ws[f'{get_column_letter(col + kk)}{row_start - 1 + row + h}'].value = npa

                        ws[f'{get_column_letter(col + 1 + kk)}{row_start}'].value = 'Point'
                        ws[f'{get_column_letter(col + 1 + kk)}{row_start + row + h}'].value = point + 1

                        ws[f'{get_column_letter(col + 1 + kk)}{row_start - 1 + row + h}'].value = point

                        ws[f'{get_column_letter(col + 2 + kk)}{row_start}'].value = 'Район'
                        ws[f'{get_column_letter(col + 2 + kk)}{row_start + row + h}'].value = name

                        ws[f'{get_column_letter(col + 2 + kk)}{row_start - 1 + row + h}'].value = name

                        ws[f'{get_column_letter(col + 3 + kk)}{row_start}'].value = f'Pнаг {name}'
                        ws[f'{get_column_letter(col + 3 + kk)}{row_start + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 3 + kk)}{row_start - 1 + row + h}'].value = pn

                        ws[f'{get_column_letter(col + 4 + kk)}{row_start}'].value = f'Pпотр {name}'
                        ws[f'{get_column_letter(col + 4 + kk)}{row_start + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 4 + kk)}{row_start - 1 + row + h}'].value = pop

                        ws[f'{get_column_letter(col + 5 + kk)}{row_start}'].value = f'Pген {name}'
                        ws[f'{get_column_letter(col + 5 + kk)}{row_start + row + h}'].value = pg

                        ws[f'{get_column_letter(col + 5 + kk)}{row_start - 1 + row + h}'].value = pg
                        kk += 5
                col += 1

    print(f"Выгружены данные из файла: smzu_mega_XML_UR_MDP.mptsmz")

    ws.title = "Потреб. районов"
    ws1 = wb.create_sheet("Графики пот.рай.")
    ws2 = wb.create_sheet("Генерация блоков")
    ws3 = wb.create_sheet("Генерация район.")
    k1 = 3
    k2 = 5
    y = 0
    u = 0
    t = 0
    for _ in range(0, 20):
        ch1 = ScatterChart()
        xvalues_ = Reference(ws,
                             min_col=2,
                             min_row=4,
                             max_row=51)

        values_ = Reference(ws,
                            min_col=k1,
                            min_row=3,
                            max_row=51)

        xvalues_2 = Reference(ws,
                              min_col=2,
                              min_row=56,
                              max_row=103)

        values_2 = Reference(ws,
                             min_col=k2,
                             min_row=55,
                             max_row=103)

        series_1 = Series(values_, xvalues_, title_from_data=True)
        series_2 = Series(values_2, xvalues_2, title_from_data=True)

        ch1.title = f"Потребление района"
        ch1.x_axis.title = "Точка, час"  # название оси Х
        ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch1.legend.position = 'b'

        ch1.series.append(series_1)
        ch1.series.append(series_2)

        ch1.x_axis.scaling.min = 25
        ch1.x_axis.scaling.max = 48

        if y == 5:
            ws1.add_chart(ch1, f'{get_column_letter(1 + u)}{str(1 + t)}')
            u = 0
            t += 16
            y = 0
        else:
            ws1.add_chart(ch1, f'{get_column_letter(1 + u)}{str(1 + t)}')
            u += 10

        y += 1

        k1 += 3
        k2 += 6

    headerCSV = ['Num', 'Point', 'P', 'ST']
    csv_2 = pd.read_csv('pgen.csv', delimiter=';', header=None)
    print(f"Загружен файл: pgen.csv")
    csv_2.columns = headerCSV
    area22 = [j for j in range(POINT_START, POINT_END)]
    row = 3
    ws2[f'{get_column_letter(1)}{2}'].value = 'Point/Num'
    for area in area22:
        d2 = csv_2[csv_2.Point == int(area)].to_dict()
        col = 1
        for i in d2['Num']:
            col += 1
            ws2[f'{get_column_letter(col)}{2}'].value = d2['Num'][i]
            ws2[f'{get_column_letter(col)}{row}'].value = d2['P'][i]
            ws2[f'{get_column_letter(1)}{row}'].value = d2['Point'][i]
        row += 1
    print(f"Выгружены данные по генерации из файла: pgen.csv")
    load_obj.load(path_file='smzu_mega_XML_UR_MDP.mptsmz', name_shabl_russian='')
    gen_table = RASTR.Tables("Generator")
    print(f"Загружен файл: smzu_mega_XML_UR_MDP.mptsmz")
    row = 0
    ws2[f"A1"].value = 'ПДГ - точка'
    ws2[f"A28"].value = 'БАРС-МДП'
    max_col = ws2.max_column
    for point in range(POINT_START, POINT_END):
        RASTR.ReadPnt(point)
        RASTR.rgm("")
        ws2[f"A{int(point) + 6}"] = point
        for i in range(2, max_col + 1):
            num_ws2 = ws2[f'{get_column_letter(i)}{2}'].value
            gen_table.SetSel(f"ID_GenBars={num_ws2}")
            j = gen_table.FindNextSel(-1)
            if j != (-1):
                ws2[f'{get_column_letter(i)}{row + 31}'].value = gen_table.Cols("P").Z(j)
                ws2[f'{get_column_letter(i)}{30}'].value = gen_table.Cols("ID_GenBars").Z(j)
                ws2[f'{get_column_letter(i)}{29}'].value = gen_table.Cols("Name").Z(j)
            else:
                ws2[f'{get_column_letter(i)}{30}'].value = num_ws2
                ws2[f'{get_column_letter(i)}{29}'].value = 'Не найден!'
            gen_table.SetSel("")
        row += 1
    load_obj.load(path_file=nameFileMPT, name_shabl_russian='')
    print(f"Загружен файл: {nameFileMPT}")
    gen_table = RASTR.Tables("Generator")
    block_table = RASTR.Tables("NBlock")
    for i in range(2, max_col + 1):
        num_agr = ws2[f'{get_column_letter(i)}{2}'].value
        block_table.SetSel(f"Num={num_agr}")
        j = block_table.FindNextSel(-1)
        if j != (-1):
            agrNum = block_table.Cols("AgrNum").Z(j)
            gen_table.SetSel(f"Num={agrNum}")
            j2 = gen_table.FindNextSel(-1)
            if j2 != (-1):
                ws2[f'{get_column_letter(i)}{1}'].value = gen_table.Cols("Name").Z(j2)
            else:
                ws2[f'{get_column_letter(i)}{1}'].value = 'Не найден в Generator!'
        else:
            ws2[f'{get_column_letter(i)}{1}'].value = 'Не найден в NBlock!'
    print(f"Выгружены данные генерации из таблицы Блоки и Генераторы УР.")

    mpt_bars = {
        15: "ЯРОСЛАВЛЬ",
        11: "ТУЛА",
        17: "ТВЕРЬ",
        28: "ТАМБОВ без ПС 220 кВ Давыдовская",
        18: "СМОЛЕНСК",
        99: "Смежные ЭС",
        31: "РЯЗАНЬ",
        96: "Потребление ПС 220 кВ Метзавод",
        98: "Потребление ПС 220 кВ Давыдовская",
        21: "ОРЕЛ",
        10: "МОСКВА",
        24: "ЛИПЕЦК",
        23: "КУРСК",
        16: "КОСТРОМА",
        32: "КАЛУГА без ПС 220 кВ Метзавод",
        13: "ИВАНОВО",
        26: "ВОРОНЕЖ",
        29: "ВОЛОГДА",
        14: "ВЛАДИМИР",
        19: "БРЯНСК",
        27: "БЕЛГОРОД",
    }

    # SMZU_Bars = {
    #     999: "ЯРОСЛАВЛЬ",
    #     998: "ТУЛА",
    #     997: "ТВЕРЬ",
    #     996: "ТАМБОВ (без учета 5980 ПС 220 кВ Давыдовская)",
    #     995: "СМОЛЕНСК",
    #     994: "РЯЗАНЬ",
    #     742: "Потребление ПС 220 кВ Метзавод",
    #     980: "Потребление ПС 220 кВ Давыдовская",
    #     993: "ОРЕЛ",
    #     992: "МОСКВА",
    #     991: "ЛИПЕЦК",
    #     990: "КУРСК",
    #     989: "КОСТРОМА",
    #     988: "КАЛУГА (без учета 5742 ПС 220 кВ Метзавод)",
    #     987: "ИВАНОВО",
    #     986: "ВОРОНЕЖ",
    #     985: "ВОЛОГДА",
    #     984: "ВЛАДИМИР",
    #     983: "БРЯНСК",
    #     982: "БЕЛГОРОД",
    # }

    filling_gen_area = {
        15: 999,
        11: 998,
        17: 997,
        28: 996,
        18: 995,
        99: 999999,
        31: 994,
        96: 742,
        98: 980,
        21: 993,
        10: 992,
        24: 991,
        23: 990,
        16: 989,
        32: 988,
        13: 987,
        26: 986,
        29: 985,
        14: 984,
        19: 983,
        27: 982,

    }

    rastr_mpt = Dispatch('Astra.Rastr')
    rastr_smzu = Dispatch('Astra.Rastr')

    load_obj_mpt = LoadFile(rastr_win=rastr_mpt)
    load_obj_smzu = LoadFile(rastr_win=rastr_smzu)

    load_obj_mpt.load(path_file=nameFileMPT, name_shabl_russian='')
    load_obj_smzu.load(path_file='smzu_mega_XML_UR_MDP.mptsmz', name_shabl_russian='')
    print(f"Загружены файлы: smzu_mega_XML_UR_MDP.mptsmz и {nameFileMPT}")
    area_rastr_mpt = rastr_mpt.Tables("area2")
    area_rastr_smzu = rastr_smzu.Tables("area2")
    g = 4
    for point in range(POINT_START, POINT_END):
        rastr_mpt.ReadPnt(point)
        rastr_mpt.rgm("")
        g += 1
        ws3[f'{get_column_letter(1)}{g}'].value = point
        rastr_smzu.ReadPnt(point)
        rastr_smzu.rgm("")
        k = 0
        for col, i in enumerate(mpt_bars):
            area_rastr_mpt.SetSel(f"npa={i}")
            j_mpt = area_rastr_mpt.FindNextSel(-1)
            area_rastr_smzu.SetSel(f"npa={filling_gen_area[i]}")
            j_smzu = area_rastr_smzu.FindNextSel(-1)
            if col == 0:
                ws3[f'{get_column_letter(col + 2)}{2}'].value = 'МТ БАРС'
                ws3[f'{get_column_letter(col + 3)}{2}'].value = 'БАРС-МДП'
                if j_mpt != (-1):
                    ws3[f'{get_column_letter(col + 2)}{4}'].value = f'{area_rastr_mpt.Cols("name").Z(j_mpt)} - МТ БАРС'
                    ws3[f'{get_column_letter(col + 2)}{3}'].value = f'{area_rastr_mpt.Cols("npa").Z(j_mpt)} - МТ БАРС'

                    ws3[f'{get_column_letter(col + 2)}{g}'].value = area_rastr_mpt.Cols("pg").Z(j_mpt)
                else:
                    ws3[f'{get_column_letter(col + 2)}{3}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 2)}{4}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 2)}{g}'].value = ''

                if j_smzu != (-1):
                    ws3[f'{get_column_letter(col + 3)}{4}'].value = f'{area_rastr_smzu.Cols("name").Z(j_smzu)} - БАРС-МДП'
                    ws3[f'{get_column_letter(col + 3)}{3}'].value = f'{area_rastr_smzu.Cols("npa").Z(j_smzu)} - БАРС-МДП'
                    ws3[f'{get_column_letter(col + 3)}{g}'].value = area_rastr_smzu.Cols("pg").Z(j_smzu)
                else:
                    ws3[f'{get_column_letter(col + 3)}{3}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 3)}{4}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 3)}{g}'].value = ''
            else:
                ws3[f'{get_column_letter(col + 4 + k)}{2}'].value = 'МТ БАРС'
                ws3[f'{get_column_letter(col + 5 + k)}{2}'].value = 'БАРС-МДП'
                if j_mpt != (-1):
                    ws3[f'{get_column_letter(col + 4 + k)}{4}'].value = f'{area_rastr_mpt.Cols("name").Z(j_mpt)} - МТ БАРС'
                    ws3[f'{get_column_letter(col + 4 + k)}{3}'].value = f'{area_rastr_mpt.Cols("npa").Z(j_mpt)} - МТ БАРС'
                    ws3[f'{get_column_letter(col + 4 + k)}{g}'].value = area_rastr_mpt.Cols("pg").Z(j_mpt)
                else:
                    ws3[f'{get_column_letter(col + 4 + k)}{3}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 4 + k)}{4}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 4 + k)}{g}'].value = ''

                if j_smzu != (-1):
                    ws3[f'{get_column_letter(col + 5 + k)}{4}'].value = f'{area_rastr_smzu.Cols("name").Z(j_smzu)} - БАРС-МДП'
                    ws3[f'{get_column_letter(col + 5 + k)}{3}'].value = f'{area_rastr_smzu.Cols("npa").Z(j_smzu)} - БАРС-МДП'
                    ws3[f'{get_column_letter(col + 5 + k)}{g}'].value = area_rastr_smzu.Cols("pg").Z(j_smzu)
                else:
                    ws3[f'{get_column_letter(col + 5 + k)}{3}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 5 + k)}{4}'].value = 'Не найден'
                    ws3[f'{get_column_letter(col + 5 + k)}{g}'].value = 'Не найден'

                k += 2

    k = 0
    h = 0
    w = 0
    j = 0
    for _ in filling_gen_area:
        ch1 = ScatterChart()
        xvalues_ = Reference(ws3,
                             min_col=1,
                             min_row=5,
                             max_row=ws3.max_row)

        values_ = Reference(ws3,
                            min_col=2 + k,
                            min_row=4,
                            max_row=ws3.max_row)

        xvalues_2 = Reference(ws3,
                              min_col=1,
                              min_row=5,
                              max_row=ws3.max_row)

        values_2 = Reference(ws3,
                             min_col=3 + k,
                             min_row=4,
                             max_row=ws3.max_row)

        series_1 = Series(values_, xvalues_, title_from_data=True)
        series_2 = Series(values_2, xvalues_2, title_from_data=True)
        name_title = ws3[f'{get_column_letter(2 + k)}{4}'].value
        name_title_list = name_title.split(" ")
        ch1.title = f"Территория: {name_title_list[0]}"
        ch1.x_axis.title = "Время, час"  # название оси Х
        ch1.y_axis.title = "Активная мощность, МВт"  # название оси У
        ch1.legend.position = 'b'

        ch1.series.append(series_1)
        ch1.series.append(series_2)

        ch1.x_axis.scaling.min = 25
        ch1.x_axis.scaling.max = 48
        j += 1
        if j == 6:
            w += 15
            j = 0
            h = 0
        ws3.add_chart(ch1, f'{get_column_letter(1 + h)}{str(34 + l)}')
        h += 10
        k += 3
    print("Выгружены данные генерации по трерриториям и построены графики.")
    wb.save('Сравнение ген. и потр. по территориям.xlsx')
    print("Сохранен файл Excel: Сравнение ген. и потр. по территориям.xlsx")


name_file_mpt_ = input(" Введите дату МТ Барс (180122-17.mpt) по формату (пример: 180122): ")
start_time = time.time()
if name_file_mpt_ != "":
    main(name_file_mpt_bars=name_file_mpt_)
else:
    print('Выход. Вы не указали дату!')

print(f'Время работы макроса:')
print(f"--- {changing_number_of_semicolons((time.time() - start_time), digits=0)} сек. ---")
print("---------------------")
print("The End.")
print("---------------------")
print("\n")
print("---------------------")
input("Нажмите Enter!")
